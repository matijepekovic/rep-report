"""
UNDISPUTED Leaderboard Mailer.

Pulls the UNDISPUTED view's DATA (not PDF) from Tableau, injects it into the
STATS leaderboard page, screenshots the rendered board with headless Chromium,
and emails the PNG.

Flow:
  1. Sign in to Tableau (same PAT secrets as pull_and_email.py).
  2. Find the saved custom view, resolve its underlying view.
  3. GET /views/{id}/data  -> summary data as CSV.
  4. Map CSV columns -> leaderboard fields (wide or Measure Names/Values long
     format both supported). CSV headers are always logged.
  5. Write data.json into the STATS checkout, serve it, screenshot 1920x1080
     (viewport auto-grows so every rep row is visible).
  6. Email the PNG. DRY_RUN=1 skips the email and just leaves artifacts.
"""

import csv
import io
import json
import os
import re
import smtplib
import subprocess
import sys
import time
from datetime import datetime, timedelta
from email.message import EmailMessage
from pathlib import Path

import requests

# ============ TABLEAU CONFIG ============

SERVER = os.getenv("TABLEAU_SERVER", "https://10ay.online.tableau.com")
SITE_CONTENT_URL = os.getenv("TABLEAU_SITE", "dabella")
API_VERSION = "3.22"

PAT_NAME = os.getenv("TABLEAU_PAT_NAME", "undisputed")
PAT_SECRET = os.getenv("TABLEAU_PAT_SECRET", "PASTE_TOKEN_SECRET_HERE")

CUSTOM_VIEW_NAME = os.getenv("CUSTOM_VIEW_NAME", "UNDISPUTED")

# The dashboard sheet holding the aggregated per-rep table. The bot pulls
# THIS sheet's data — the lead-level details sheet is irrelevant.
TARGET_SHEET = os.getenv("TARGET_SHEET", "Sales Rep Totals")

# Optional view filters for the data query, e.g. "Region=West;Week=Last"
# (applied as vf_<field>=<value> query params).
VIZ_FILTERS = os.getenv("TABLEAU_VIZ_FILTERS", "")

# ============ BOARD / RENDER CONFIG ============

STATS_DIR = Path(os.getenv("STATS_DIR", "stats-board"))  # checkout of matijepekovic/STATS
SHOT_WIDTH = int(os.getenv("SHOT_WIDTH", "1920"))
SHOT_HEIGHT = int(os.getenv("SHOT_HEIGHT", "1080"))
OUT_PNG = Path(os.getenv("OUT_PNG", "board.png"))

# Week range label; override with REPORT_RANGE="MAY 19 - MAY 25, 2025"
REPORT_RANGE = os.getenv("REPORT_RANGE", "")

DRY_RUN = os.getenv("DRY_RUN", "0").lower() in ("1", "true", "yes")

# ============ EMAIL CONFIG ============

SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = 587
SMTP_USER = os.getenv("SMTP_USER", "you@gmail.com")
SMTP_PASS = os.getenv("SMTP_PASS", "APP_PASSWORD_HERE")

EMAIL_TO = os.getenv("EMAIL_TO", "you@yourcompany.com").split(",")
EMAIL_SUBJECT = os.getenv("EMAIL_SUBJECT", "UNDISPUTED Leaderboard")

# ============ CSV -> BOARD FIELD MAPPING ============
# Aliases are matched against normalized headers (lowercase, alphanumerics only).
# Adjust the alias lists if the first run's logged headers differ.

FIELD_ALIASES = {
    "issuedLeads":    ["issuedleads", "leadsissued", "issued"],
    "pitchedLeads":   ["pitchedleads", "leadspitched", "pitched"],
    "pitchRate":      ["pitchedrate", "pitchrate", "pitchpct", "pitchpercent"],
    "soldLeads":      ["soldleads", "leadssold", "sold"],
    "closeRate":      ["closerate", "closingrate", "closepct", "closepercent"],
    "grossSplit":     ["grosssplit", "grosssales", "grossvolume", "gross"],
    "netSplit":       ["netsplit", "netsales", "netvolume"],
    "pendingSplit":   ["pendingsplit", "pending"],
    "dpl":            ["dollarsperlead", "dpl", "perlead"],
    "salesRetention": ["salesretention", "retentionrate", "retention"],
    "avgGrossPerRep": ["avggrosssaleperrep", "avggrossperrep", "averagegrossperrep", "avggross"],
    "avgNetPerRep":   ["avgnetsaleperrep", "avgnetperrep", "averagenetperrep", "avgnet"],
}
# (alias, field) pairs, longest alias first, so specific names win
ALIAS_INDEX = sorted(
    ((a, f) for f, aliases in FIELD_ALIASES.items() for a in aliases),
    key=lambda p: -len(p[0]))

REP_ALIASES = ["srname", "rep", "repname", "salesrep", "salesperson",
               "employee", "assignedrep"]
BRANCH_ALIASES = ["branchnew", "branch", "office", "location", "market"]
MEASURE_NAME_COLS = ["measurenames", "measure"]
MEASURE_VALUE_COLS = ["measurevalues", "value"]
TOTAL_ROW_NAMES = {"grand total", "total", "totals", "team total", "team totals"}

PCT_FIELDS = {"pitchRate", "closeRate", "salesRetention"}
# aggregated as mean over rows; everything else sums
MEAN_FIELDS = PCT_FIELDS | {"dpl", "avgGrossPerRep", "avgNetPerRep"}


def log(msg):
    print(f"[{datetime.now():%H:%M:%S}] {msg}", flush=True)


def norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def clean_number(raw, pct=False):
    """'$48,750' -> 48750.0 ; '82%' -> 82.0 ; '' -> None."""
    if raw is None:
        return None
    s = str(raw).strip().replace("$", "").replace(",", "").replace("%", "")
    if s in ("", "-", "—", "n/a", "null", "None"):
        return None
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if neg else v


# ---------------------------------------------------------------- Tableau

def tableau_signin():
    api_ver = API_VERSION
    try:
        info = requests.get(f"{SERVER}/api/{API_VERSION}/serverinfo",
                            headers={"Accept": "application/json"}, timeout=30).json()
        api_ver = info["serverInfo"]["restApiVersion"]
    except Exception:
        pass
    log(f"Using API version {api_ver}")
    base = f"{SERVER}/api/{api_ver}"
    s = requests.Session()
    s.headers["Accept"] = "application/json"
    r = s.post(f"{base}/auth/signin", json={
        "credentials": {
            "personalAccessTokenName": PAT_NAME,
            "personalAccessTokenSecret": PAT_SECRET,
            "site": {"contentUrl": SITE_CONTENT_URL},
        }
    })
    if r.status_code != 200:
        log(f"Sign-in failed ({r.status_code}): {r.text[:300]}")
        sys.exit(1)
    creds = r.json()["credentials"]
    s.headers["X-Tableau-Auth"] = creds["token"]
    log("Signed in.")
    return s, base, creds["site"]["id"]


def fetch_view_csv(s, base, site_id) -> str:
    # Locate the saved custom view (same as the PDF bot), then use its
    # underlying view for the data query.
    r = s.get(f"{base}/sites/{site_id}/customviews", params={"pageSize": 1000})
    r.raise_for_status()
    cvs = r.json().get("customViews", {}).get("customView", [])
    cv = next((c for c in cvs
               if c["name"].strip().lower() == CUSTOM_VIEW_NAME.strip().lower()), None)
    if not cv:
        log(f"Custom view '{CUSTOM_VIEW_NAME}' not found. Visible custom views:")
        for c in cvs[:50]:
            log(f"  - {c['name']}")
        sys.exit(1)
    view_id = cv.get("view", {}).get("id")
    log(f"Custom view '{cv['name']}' -> underlying view {view_id}")
    if not view_id:
        log("Custom view carries no underlying view id; cannot query data.")
        sys.exit(1)

    params = {"maxAge": "1"}
    for pair in [p for p in VIZ_FILTERS.split(";") if "=" in p]:
        k, v = pair.split("=", 1)
        params[f"vf_{k.strip()}"] = v.strip()

    # Numbers come from the Rep Totals table AS DISPLAYED, via the crosstab
    # endpoint — one row per rep, every column aggregated by Tableau. The
    # default view is unfiltered (whole company), so the custom view's DATA
    # export (which does honor the saved filters) is used only to learn
    # WHICH reps belong on the board — never for numbers.
    allowed = set()
    r = s.get(f"{base}/sites/{site_id}/customviews/{cv['id']}/data",
              params={"maxAge": "1"})
    if r.status_code == 200:
        allowed = rep_names_in_csv(r.content.decode("utf-8-sig"))
        log(f"Custom view scope: {len(allowed)} reps -> {sorted(allowed)}")
    else:
        log(f"Custom-view data unavailable ({r.status_code}); board will "
            "include every rep in the crosstab.")

    r = s.get(f"{base}/sites/{site_id}/views/{view_id}/crosstab/excel",
              params={**params, "maxAge": "1"})
    if r.status_code != 200:
        log(f"Crosstab pull failed ({r.status_code}): {r.text[:300]}")
        sys.exit(1)
    log(f"Got crosstab xlsx ({len(r.content)//1024} KB)")
    return crosstab_to_csv(r.content), allowed


def rep_names_in_csv(csv_text: str) -> set:
    reader = csv.DictReader(io.StringIO(csv_text))
    rep_col = next((h for h in (reader.fieldnames or [])
                    if norm(h) in REP_ALIASES), None)
    if not rep_col:
        return set()
    return {(row.get(rep_col) or "").strip()
            for row in reader if (row.get(rep_col) or "").strip()}


def crosstab_to_csv(xlsx_bytes: bytes) -> str:
    """Pick the aggregated totals worksheet out of the crosstab workbook
    (a dashboard exports one worksheet per table) and flatten it to CSV."""
    import io as _io
    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    log("Crosstab worksheets: " + ", ".join(f"'{w.title}'" for w in wb.worksheets))

    def sheet_rows(ws):
        return [["" if c is None else str(c) for c in row]
                for row in ws.iter_rows(values_only=True)]

    best = None
    for ws in wb.worksheets:
        rows = sheet_rows(ws)
        hdr = next((i for i, row in enumerate(rows)
                    if any(norm(c) in REP_ALIASES for c in row)), None)
        if hdr is None:
            continue
        header_norms = [norm(c) for c in rows[hdr]]
        # the aggregated table is the one carrying computed columns like DPL
        score = sum(1 for n in header_norms
                    if any(k in n for k in ("dpl", "retention", "avggross", "avgnet")))
        if best is None or score > best[0]:
            best = (score, ws.title, rows[hdr:])
    wb.close()
    if best is None:
        log("No worksheet with a rep column found in the crosstab.")
        sys.exit(1)
    log(f"Using worksheet '{best[1]}' (computed-column score {best[0]})")
    out = _io.StringIO()
    csv.writer(out).writerows(best[2])
    return out.getvalue()


# ---------------------------------------------------------------- Mapping

def match_field(header_norm):
    """Longest-alias-first substring match, so e.g. 'AGG(Pitch Rate)' or
    'SUM(Net Split)' still resolve to the right field."""
    if not header_norm:
        return None
    for alias, field in ALIAS_INDEX:
        if alias in header_norm:
            return field
    return None


def rows_to_board(csv_text: str, allowed: set = None):
    reader = csv.DictReader(io.StringIO(csv_text))
    headers = reader.fieldnames or []
    log(f"CSV headers: {headers}")
    hmap = {h: norm(h) for h in headers}

    rep_col = next((h for h, n in hmap.items() if n in REP_ALIASES), None)
    branch_col = next((h for h, n in hmap.items() if n in BRANCH_ALIASES), None)
    mn_col = next((h for h, n in hmap.items() if n in MEASURE_NAME_COLS), None)
    mv_col = next((h for h, n in hmap.items() if n in MEASURE_VALUE_COLS), None)
    if not rep_col:
        log("No rep-name column recognized. Add its header to REP_ALIASES.")
        sys.exit(1)
    long_fmt = bool(mn_col and mv_col)
    log(f"Rep column: '{rep_col}'  Branch column: '{branch_col}'  "
        f"Long format: {long_fmt}")

    # acc[name][field] = [sum, count]; the CSV may hold many rows per rep
    # (lead-level detail), so measures accumulate instead of overwrite.
    # The export can repeat the same (lead, rep, measure) row — e.g. the viz
    # fanning a lead out over both sales-rep fields — which doubles every
    # number, so rows are deduped on that key before summing.
    lead_col = next((h for h, n in hmap.items() if n == "leadid"), None)
    acc = {}
    order = []
    measure_seen = {}   # raw measure name -> (row count, mapped field)
    seen_keys = {}      # (rep, lead, measure) -> occurrences
    n_rows = n_dupes = 0
    for row in reader:
        n_rows += 1
        name = (row.get(rep_col) or "").strip()
        if not name:
            continue
        if name not in acc:
            acc[name] = {}
            order.append(name)

        def feed(field, raw):
            v = clean_number(raw, field in PCT_FIELDS)
            if v is None:
                return
            s = acc[name].setdefault(field, [0.0, 0])
            s[0] += v
            s[1] += 1

        if long_fmt:
            mname = (row.get(mn_col) or "").strip()
            field = match_field(norm(mname))
            cnt, _ = measure_seen.get(mname, (0, field))
            measure_seen[mname] = (cnt + 1, field)
            lead = (row.get(lead_col) or "").strip() if lead_col else ""
            if lead:                      # dedupe repeated lead rows exactly
                key = (name, lead, mname)
                seen_keys[key] = seen_keys.get(key, 0) + 1
                if seen_keys[key] > 1:
                    n_dupes += 1
                    continue
            if field:
                feed(field, row.get(mv_col))
        else:
            for h, n in hmap.items():
                field = match_field(n)
                if field:
                    feed(field, row.get(h))

    if long_fmt:
        log(f"{n_rows} rows, {n_dupes} duplicate (rep,lead,measure) rows dropped.")
        if seen_keys:
            from collections import Counter
            hist = Counter(seen_keys.values())
            log(f"Rows-per-key histogram: {dict(sorted(hist.items()))}")
        log("Distinct Measure Names -> mapped field:")
        for mname, (cnt, field) in sorted(measure_seen.items()):
            log(f"  - '{mname}' x{cnt} -> {field or 'UNMAPPED'}")

    # Collapse accumulators: sums for counts/dollars, means for rates & avgs.
    # Board shows rep name only — no branch.
    reps = {}
    for name in order:
        rec = {"name": name}
        for field, (total, count) in acc[name].items():
            rec[field] = (total / count) if field in MEAN_FIELDS and count else total
        reps[name] = rec

    # Split out a totals row if Tableau included one
    totals = None
    for name in list(reps):
        if name.strip().lower() in TOTAL_ROW_NAMES:
            totals = reps.pop(name)
            totals.pop("name", None)
            totals.pop("branch", None)
            derive(totals)
            order.remove(name)

    rep_list = [reps[n] for n in order]

    # Keep only the reps the custom view shows; totals are then recomputed
    # from the kept rows (the crosstab's Grand Total spans the whole company).
    if allowed:
        allowed_lc = {a.strip().lower() for a in allowed}
        before = len(rep_list)
        rep_list = [r for r in rep_list if r["name"].strip().lower() in allowed_lc]
        log(f"Scope filter: kept {len(rep_list)} of {before} reps")
        totals = None

    if not rep_list:
        log("No rep rows parsed — check the header mapping above.")
        sys.exit(1)

    # Derived columns, formulas verified against the Tableau PDF export:
    #   Pitch Rate = Pitched/Issued        Close Rate = Sold/Issued
    #   DPL = Net/Issued                   Sales Retention = Net/Gross
    #   Avg Gross Sale per Rep = Gross/Sold   Avg Net Sale per Rep = Net/Sold
    for r in rep_list:
        derive(r)

    if totals is None:
        totals = compute_totals(rep_list)
        log("No total row in CSV; computed totals from rep rows.")

    log(f"Parsed {len(rep_list)} reps. Fields on first rep: "
        f"{sorted(k for k in rep_list[0] if k not in ('name', 'branch'))}")
    for r in rep_list[:3]:
        log(f"  sample: {r}")
    return rep_list, totals


def derive(r):
    """Fill MISSING computed columns in-place (rep row or totals row).
    Values that came straight from the Tableau sheet are never overwritten."""
    def put(key, val):
        if r.get(key) is None:
            r[key] = val
    issued, pitched = r.get("issuedLeads"), r.get("pitchedLeads")
    sold = r.get("soldLeads")
    gross, net = r.get("grossSplit"), r.get("netSplit")
    if issued and pitched is not None:
        put("pitchRate", pitched / issued)
    if issued and sold is not None:
        put("closeRate", sold / issued)
    if issued and net is not None:
        put("dpl", net / issued)
    if gross and net is not None:
        put("salesRetention", net / gross)
    if sold:
        if gross is not None:
            put("avgGrossPerRep", gross / sold)
        if net is not None:
            put("avgNetPerRep", net / sold)


def compute_totals(rep_list):
    def total(key):
        vals = [r.get(key) for r in rep_list if r.get(key) is not None]
        return sum(vals) if vals else None

    t = {k: total(k) for k in
         ("issuedLeads", "pitchedLeads", "soldLeads",
          "grossSplit", "pendingSplit", "netSplit")}
    derive(t)
    return t


def default_range():
    # The saved view covers the current month.
    return f"{datetime.now():%B %Y}".upper()


# ---------------------------------------------------------------- Render

def render_board(data: dict) -> Path:
    if not (STATS_DIR / "index.html").exists():
        log(f"STATS checkout not found at '{STATS_DIR}'.")
        sys.exit(1)
    (STATS_DIR / "data.json").write_text(json.dumps(data, indent=1))
    log(f"Wrote {STATS_DIR}/data.json")
    if DRY_RUN:
        log("DATAJSON " + json.dumps(data, separators=(",", ":")))

    server = subprocess.Popen(
        [sys.executable, "-m", "http.server", "8123", "--bind", "127.0.0.1"],
        cwd=STATS_DIR, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    try:
        time.sleep(1.0)
        from playwright.sync_api import sync_playwright
        with sync_playwright() as pw:
            exe = os.getenv("CHROMIUM_PATH")  # optional override
            browser = pw.chromium.launch(executable_path=exe) if exe else pw.chromium.launch()
            page = browser.new_page(viewport={"width": SHOT_WIDTH, "height": SHOT_HEIGHT})
            page.goto("http://127.0.0.1:8123/index.html", wait_until="networkidle")
            page.wait_for_timeout(700)
            # Grow the viewport until every rep row is visible (board scrolls
            # internally when there are more reps than fit).
            for _ in range(6):
                need = page.evaluate(
                    "()=>{const b=document.getElementById('board');"
                    "return b? b.scrollHeight-b.clientHeight : 0}")
                if need <= 2:
                    break
                h = min(page.viewport_size["height"] + need + 40, 3000)
                page.set_viewport_size({"width": SHOT_WIDTH, "height": h})
                page.wait_for_timeout(300)
            page.screenshot(path=str(OUT_PNG))
            browser.close()
    finally:
        server.terminate()
    log(f"Screenshot saved: {OUT_PNG} ({OUT_PNG.stat().st_size//1024} KB)")
    return OUT_PNG


# ---------------------------------------------------------------- Email

def send_email(png: Path):
    msg = EmailMessage()
    today = datetime.now().strftime("%b %d, %Y")
    msg["Subject"] = f"{EMAIL_SUBJECT} — {today}"
    msg["From"] = SMTP_USER
    msg["To"] = ", ".join(EMAIL_TO)
    msg.set_content(f"UNDISPUTED leaderboard for {today}. Attached.\n\n"
                    "— Automated by the rep board.")
    msg.add_attachment(png.read_bytes(), maintype="image", subtype="png",
                       filename=f"Undisputed_{datetime.now():%Y-%m-%d}.png")
    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as smtp:
        smtp.starttls()
        smtp.login(SMTP_USER, SMTP_PASS)
        smtp.send_message(msg)
    log(f"Emailed to {', '.join(EMAIL_TO)}")


# ---------------------------------------------------------------- Main

if __name__ == "__main__":
    s, base, site_id = tableau_signin()
    csv_text, allowed = fetch_view_csv(s, base, site_id)
    Path("view_data.csv").write_text(csv_text)
    s.post(f"{base}/auth/signout")

    rep_list, totals = rows_to_board(csv_text, allowed)
    data = {
        "dateRange": REPORT_RANGE or default_range(),
        "reps": rep_list,
        "totals": totals,
    }
    png = render_board(data)

    if DRY_RUN:
        log("DRY_RUN set — skipping email. Inspect board.png / view_data.csv.")
    else:
        send_email(png)
    log("Done.")
